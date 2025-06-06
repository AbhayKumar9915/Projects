import openpyxl

file = "C:\\Users\\ABHAY\\Test.xlsx"

wb = openpyxl.load_workbook(file)
sheet = wb['Sheet1']

print('***Number of Rows and Column***')
print(sheet.max_row)
print(sheet.max_column)

print('***Desired Value***')
print(sheet.cell(1, 2).value)

print("***All Column Names***")
for i in range(1, sheet.max_column + 1):
    cell_obj = sheet.cell(row=1, column=i)
    print(cell_obj.value)

print('***First Column Values')
for i in range(1, sheet.max_row + 1):
    cell_obj = sheet.cell(row=i, column=1)
    print(cell_obj.value)
