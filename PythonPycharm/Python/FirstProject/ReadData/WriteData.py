import openpyxl

file = "C:\\Users\\ABHAY\\Test.xlsx"

wb = openpyxl.load_workbook(file)
sheet = wb['Sheet1']

# sheet object cell method
c1 = sheet.cell(5, 1)
c1.value = 'Abhay'
c2 = sheet.cell(5, 2)
c2.value = "Verma"
# cell object can call by its name
c3 = sheet['A6']
c3.value = 'Ruchi'
c4 = sheet['B6']
c4.value = 'Verma'

wb.save(file)